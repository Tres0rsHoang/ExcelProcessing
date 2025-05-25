import pandas as pd
import sys
from openpyxl import load_workbook
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QFont
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QLabel, 
    QPushButton, QFileDialog, QVBoxLayout,
    QWidget, QApplication, QHBoxLayout,
    QLineEdit, QGridLayout, QScrollArea,
)

def searchData(file_path, search_string):
    data = pd.read_excel(file_path, skiprows=8)
    result = []
   
    for i in range(len(data)):
        if search_string.lower() in data.iloc[i]['CÂU HỎI'].lower():
            question = data.iloc[i]['CÂU HỎI']
            answerIndex = data.iloc[i]['ĐÁP ÁN ĐÚNG'] + 1
            answer = data.iloc[i, answerIndex]
            result.append([question, str(answer), str(data.iloc[i]['ĐÁP ÁN ĐÚNG'])])
            
    return result

class RemoveDuplicateWindow(QMainWindow):
    def __init__(self, file_path=""):
        super().__init__()
        self.file_path = file_path
        self.folder_path = ""
        self.setWindowTitle("Remove Duplicates - Python")
        self.setFixedSize(600, 100)
        self.center()
                
        self.folder_label = QLabel("Select an output path", self)
        self.folder_button = QPushButton("Browse folder", self)
        self.folder_button.clicked.connect(self.pick_folder)
        
        self.cancel_button = QPushButton("Cancel", self)
        self.remove_button = QPushButton("Remove", self)
        self.cancel_button.clicked.connect(lambda: self.close())
        self.remove_button.clicked.connect(self.remove_duplicates)
        
        self.initUI()
    
    def initUI(self):
        central_widget = QWidget()
        self.setStyleSheet("background-color: #323232;")
        self.setCentralWidget(central_widget)
        
        folder_selection_layout = QHBoxLayout()
        self.folder_label.setStyleSheet("background-color: #151E29; border-radius: 8px; padding: 10px; color: white;")
        self.folder_button.setStyleSheet("border-radius: 8px; padding: 10px; color: white; background-color:#656565")
        folder_selection_layout.addWidget(self.folder_label, 7)
        folder_selection_layout.addWidget(self.folder_button, 3)
        
        confirm_layout = QHBoxLayout()
        self.cancel_button.setStyleSheet("border-radius: 8px; padding: 10px; color: white; background-color:#656565")
        self.remove_button.setStyleSheet("border-radius: 8px; padding: 10px; color: white; background-color:#FF3333")
        confirm_layout.addWidget(self.cancel_button, 1)
        confirm_layout.addWidget(self.remove_button, 1)
        
        layout = QVBoxLayout()
        layout.setAlignment(Qt.AlignTop | Qt.AlignLeft)
        layout.addLayout(folder_selection_layout)
        layout.addLayout(confirm_layout)
        
        central_widget.setLayout(layout)
    
    def remove_duplicates(self):
        if self.folder_path == "" or self.file_path == "": return
        
        df = pd.read_excel(self.file_path, skiprows=8)
        col_b = df.columns[1]  
        df_unique = df.drop_duplicates(subset=[col_b], keep='first')

        rows_to_keep = df_unique.index.tolist()
        
        wb = load_workbook(self.file_path)
        ws = wb.active

        all_data_rows = range(8, ws.max_row + 1)
        for row in reversed(all_data_rows):
            if (row - 8) not in rows_to_keep:
                ws.delete_rows(row)

        wb.save(f"{self.folder_path}/RemoveDuplicatesResult.xlsx")
        
        self.close()
        
    def pick_folder(self):
        folder_path = QFileDialog.getExistingDirectory(self, "Select Folder")
        if folder_path:
            self.folder_label.setText(f"{folder_path}")
            self.folder_path = folder_path
        
    def center(self):
        screen_geometry = QApplication.desktop().screenGeometry()
        window_geometry = self.frameGeometry()
        center_point = screen_geometry.center()
        window_geometry.moveCenter(center_point)
        self.move(window_geometry.topLeft())

class MainWindow(QMainWindow):
    
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Data Crawler - Python")
        self.resize(800, 600)
        self.center()
        self.remove_duplicates_window = None
        self.file_path = ""
      
        self.file_label = QLabel("Select a file to start", self)
        self.file_button = QPushButton("Browse file", self)
        self.file_button.clicked.connect(self.pick_file)
        self.remove_duplicates_button = QPushButton("Remove duplicates", self)
        self.remove_duplicates_button.clicked.connect(self.open_remove_duplicates_window)
        
        self.input_field = QLineEdit(self)
        self.input_field.setPlaceholderText("Enter search string")
        self.search_button = QPushButton("Search", self)
        self.search_button.clicked.connect(self.search)
        self.initUI()
        
    def initUI(self):
        central_widget = QWidget()
        self.setStyleSheet("background-color: #323232;")
        self.setCentralWidget(central_widget)
        
        # File selection layout
        file_selection_layout = QHBoxLayout()
        self.file_label.setStyleSheet("background-color: #151E29; border-radius: 8px; padding: 10px; color: white;")
        self.file_button.setStyleSheet("border-radius: 8px; padding: 10px; color: white; background-color:#656565")
        self.remove_duplicates_button.setStyleSheet("border-radius: 8px; padding: 10px; color: white; background-color:#656565")
        file_selection_layout.addWidget(self.file_label, 7)
        file_selection_layout.addWidget(self.file_button, 1)
        file_selection_layout.addWidget(self.remove_duplicates_button, 1)
        
        # Search layout
        search_layout = QHBoxLayout()
        self.input_field.setStyleSheet("background-color: #141D27; font-size:16pt; border-radius: 8px; padding: 10px; color: white;")
        self.search_button.setStyleSheet("width: 100px; border-radius: 8px; font-size:16pt; padding: 10px; color: white; background-color:#656565")
        search_layout.addWidget(self.input_field, 7)
        search_layout.addWidget(self.search_button, 3)
        
        # Result table layout
        self.table_layout = QGridLayout()
        self.table_layout.setAlignment(Qt.AlignTop | Qt.AlignLeft)
        self.table_layout.setColumnStretch(0, 20)
        self.table_layout.setColumnStretch(1, 1)
        self.table_layout.setColumnStretch(2, 20)
        scroll_content = QWidget()
        scroll_content.setLayout(self.table_layout)
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setWidget(scroll_content)
        scroll_area.setStyleSheet("""
            QScrollBar:vertical {
                background: #2e2e2e;
                width: 6px;
                margin: 0px 0px 0px 0px;
                border-radius: 6px;
            }
            QScrollBar::handle:vertical {
                background: #606060;
                min-height: 20px;
                border-radius: 6px;
            }
            QScrollBar::handle:vertical:hover {
                background: #909090;
            }
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
                height: 0px;
                background: none;
            }
            QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical {
                background: none;
            }
        """)
        
        layout = QVBoxLayout()
        layout.setAlignment(Qt.AlignTop | Qt.AlignLeft)
        layout.addLayout(file_selection_layout)
        layout.addLayout(search_layout)
        layout.addWidget(scroll_area)

        central_widget.setLayout(layout)
    
    def pick_file(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "Open File", "", "Excel (*.xls *.xlsx)")
        if file_path:
            self.file_label.setText(f"{file_path}")
            self.file_path = file_path
            
    def clear_table(self):
        while self.table_layout.count() :
            item = self.table_layout.takeAt(0)
            widget = item.widget()
            if widget:
                widget.setParent(None)
                widget.deleteLater()
            elif item.layout() is not None:
                self.clear_table(item.layout())
            
    def search(self):
        self.clear_table()
        if (self.file_path == "" or self.input_field.text() == ""): return
        result = searchData(self.file_path, self.input_field.text())    
        
        questionTile = QLabel("Question")
        answerTile = QLabel("Answer")
        indexTile = QLabel("Index")
        
        questionTile.setStyleSheet("font-weight: bold; color: white; font-size: 16pt;")
        answerTile.setStyleSheet("font-weight: bold; color: white; font-size: 16pt;")
        indexTile.setStyleSheet("font-weight: bold; color: white; font-size: 16pt;")
        
        self.table_layout.addWidget(questionTile, 0, 0, alignment=Qt.AlignCenter)
        self.table_layout.addWidget(indexTile, 0, 1, alignment=Qt.AlignCenter)
        self.table_layout.addWidget(answerTile, 0, 2, alignment=Qt.AlignCenter)
        
        for i in range(len(result)):
            question = QLabel(result[i][0])
            answer = QLabel(result[i][1])
            index = QLabel(result[i][2])
            
            question.setStyleSheet("background-color: #2F2F2F; border-radius: 8px; font-size: 16pt; padding: 4px; color: white;")
            answer.setStyleSheet("background-color: #607d29; border-radius: 8px; font-size: 16pt; padding: 4px; color: white;")
            index.setStyleSheet("background-color: #2F2F2F; border-radius: 8px; font-size: 16pt; padding: 4px; color: #B1C34E;")
            
            scroll_area = QScrollArea()
            scroll_area.setWidgetResizable(True)
            scroll_area.setWidget(question)
            scroll_area.setStyleSheet("""
                QScrollArea {
                    background-color: #2F2F2F; 
                    border-radius: 8px;
                }                                      
                QScrollBar:vertical {
                    background: #2e2e2e;
                    width: 6px;
                    margin: 0px 0px 0px 0px;
                    border-radius: 6px;
                }
                QScrollBar::handle:vertical {
                    background: #606060;
                    min-height: 20px;
                    border-radius: 6px;
                }
                QScrollBar::handle:vertical:hover {
                    background: #909090;
                }
                QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
                    height: 0px;
                    background: none;
                }
                QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical {
                    background: none;
                }
            """)
            
            question.setWordWrap(True)
            answer.setWordWrap(True)
            index.setWordWrap(True)
            
            self.table_layout.addWidget(scroll_area, i+1, 0)
            self.table_layout.addWidget(index, i+1, 1, alignment=Qt.AlignCenter)
            self.table_layout.addWidget(answer, i+1, 2)
    
    def open_remove_duplicates_window(self):
        if self.file_path == "": return
        
        if self.remove_duplicates_window is None:
            self.remove_duplicates_window = RemoveDuplicateWindow(file_path=self.file_path)
        self.remove_duplicates_window.show()
        
    def center(self):
        screen_geometry = QApplication.desktop().screenGeometry()
        window_geometry = self.frameGeometry()
        center_point = screen_geometry.center()
        window_geometry.moveCenter(center_point)
        self.move(window_geometry.topLeft())    
        
def main():
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())

if __name__ == "__main__":
    main()
